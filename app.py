import streamlit as st
import pandas as pd
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="MCF Admission Auditor", layout="wide")

st.title("📊 MCF Admission Analyzer + Audit System")
st.caption("Created by CA Mahesh Godase")

uploaded_file = st.file_uploader("📂 Upload Admission File", type=["xlsx"])

if uploaded_file is not None:
    try:
        df_raw = pd.read_excel(uploaded_file)
        df_raw.columns = df_raw.columns.astype(str).str.strip()
        df_raw_display = df_raw.fillna("")
        st.subheader("🔍 Raw Data Preview")
        st.dataframe(df_raw_display)

        def find_col(keys):
            for col in df_raw.columns:
                for k in keys:
                    if k in col.lower():
                        return col
            return None

        emp_col = find_col(["employee", "staff", "counsellor"])
        camp_col = find_col(["camp"])

        if emp_col is None or camp_col is None:
            st.error("❌ Required columns not found")
            st.stop()

        df_calc = df_raw[[emp_col, camp_col]].dropna()
        df_calc[emp_col] = df_calc[emp_col].astype(str).str.strip().str.upper()
        df_calc[camp_col] = df_calc[camp_col].astype(str).str.strip()

        pivot = pd.pivot_table(df_calc, index=emp_col, columns=camp_col, aggfunc="size", fill_value=0)
        pivot = pivot.reset_index()
        pivot.rename(columns={emp_col: "Employee Name"}, inplace=True)
        pivot["Total"] = pivot.iloc[:, 1:].sum(axis=1)
        total_row = pd.DataFrame(pivot.iloc[:, 1:].sum()).T
        total_row.insert(0, "Employee Name", "Total")
        final_df = pd.concat([pivot, total_row], ignore_index=True)

        st.subheader("📋 Final Report")
        st.dataframe(final_df, use_container_width=True)

        def auto_width(ws):
            for col_cells in ws.iter_cols():
                max_len = max((len(str(c.value)) for c in col_cells if c.value), default=0)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 3

        # Border & Center Alignment
        thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))
        center = Alignment(horizontal="center", vertical="center")
        bold_font = Font(bold=True)

        def apply_table_style(ws, start_row=1, start_col=1, max_row=None, max_col=None):
            for r in range(start_row, max_row+1):
                for c in range(start_col, max_col+1):
                    ws.cell(row=r, column=c).border = thin
                    ws.cell(row=r, column=c).alignment = center

        def to_excel(df, raw_df):
            wb = Workbook()
            header_fill = PatternFill(start_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            title_font = Font(size=16, bold=True)

            # ===== ADMISSION REPORT =====
            ws = wb.active
            ws.title = "Admission Report"
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(df.columns))
            ws.cell(1, 1).value = "MCF Admission MIS Report"
            ws.cell(1, 1).font = title_font
            ws.cell(1, 1).alignment = center
            start_row = 4
            for col_num, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=start_row, column=col_num, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            for r, row in enumerate(df.values, start_row + 1):
                ws.append(list(row))
            apply_table_style(ws, start_row=start_row, start_col=1, max_row=ws.max_row, max_col=len(df.columns))
            ws.freeze_panes = "A5"
            auto_width(ws)

            # ===== FEES ANALYSIS =====
            ws8 = wb.create_sheet("Fees Analysis", 1)
            fee_col = None
            for col in raw_df.columns:
                if "fee" in col.lower() or "amount" in col.lower():
                    fee_col = col
            if fee_col:
                temp_fee = raw_df.copy()
                temp_fee[fee_col] = pd.to_numeric(temp_fee[fee_col], errors="coerce").fillna(0)
                summary = temp_fee.groupby(camp_col)[fee_col].sum().reset_index()
                summary = summary.sort_values(by=fee_col, ascending=False)
                summary["%"] = round(summary[fee_col]/summary[fee_col].sum()*100,2)
                summary["Rank"] = range(1,len(summary)+1)
                ws8.append(["Rank","Camp","Fees","%"])
                for row in summary.values:
                    ws8.append(list(row))
                # Format % column
                for r in range(2, ws8.max_row+1):
                    ws8.cell(row=r, column=4).number_format = '0.00%'
                for c in range(1,5):
                    ws8.cell(row=1,column=c).font = bold_font
                    ws8.cell(row=1,column=c).fill = PatternFill(start_color="4F81BD", fill_type="solid")
                    ws8.cell(row=1,column=c).alignment = center
                apply_table_style(ws8, start_row=1, start_col=1, max_row=ws8.max_row, max_col=4)
                # Pie Chart
                pie = PieChart()
                pie.title = "Camp-wise Fees Distribution"
                data = Reference(ws8, min_col=3, min_row=1, max_row=ws8.max_row)
                labels = Reference(ws8, min_col=2, min_row=2, max_row=ws8.max_row)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.width = 12
                pie.height = 10
                ws8.add_chart(pie,"F2")
            else:
                ws8.append(["No Fees Column Found"])
            auto_width(ws8)

            # ===== DASHBOARD =====
            ws2 = wb.create_sheet("Dashboard")
            chart = BarChart()
            data = Reference(ws, min_col=len(df.columns), min_row=4, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=5, max_row=ws.max_row-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            ws2.add_chart(chart, "A1")

            # ===== CAMP ANALYSIS =====
            ws3 = wb.create_sheet("Camp Analysis")
            ws3.append(["Rank","Camp","Total","%"])
            for col_num in range(1,5):
                cell = ws3.cell(row=1, column=col_num)
                cell.font = bold_font
                cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
                cell.alignment = center
            total = df.iloc[-1]["Total"]
            camp_list = df.columns[1:-1]
            camp_totals = [df.iloc[-1][c] for c in camp_list]
            camp_ranking = sorted(zip(camp_list,camp_totals), key=lambda x:x[1], reverse=True)
            for idx,(camp,val) in enumerate(camp_ranking,1):
                ws3.append([idx,camp,val,val/total])
            for r in range(2, ws3.max_row+1):
                ws3.cell(row=r, column=4).number_format = '0.00%'
            apply_table_style(ws3, start_row=1, start_col=1, max_row=ws3.max_row, max_col=4)
            # Pie chart
            pie = PieChart()
            pie.title = "Camp-wise Distribution"
            data = Reference(ws3, min_col=3, min_row=1, max_row=ws3.max_row)
            labels = Reference(ws3, min_col=2, min_row=2, max_row=ws3.max_row)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.width=12; pie.height=10
            ws3.add_chart(pie,"F2")
            # Bar chart
            bar = BarChart()
            bar.title = "Camp-wise Total"
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(labels)
            bar.width=16; bar.height=10
            ws3.add_chart(bar,"F18")

            # ===== TOP PERFORMERS =====
            ws4 = wb.create_sheet("Top Performers")
            temp = df.iloc[:-1].sort_values(by="Total", ascending=False)
            ws4.append(["Rank","Employee","Total"])
            for i,row in enumerate(temp.values,1):
                ws4.append([i,row[0],row[-1]])
            apply_table_style(ws4,start_row=1,start_col=1,max_row=ws4.max_row,max_col=3)

            # ===== RAW DATA =====
            ws6 = wb.create_sheet("Raw Data")
            ws6.append(list(raw_df.columns))
            for row in raw_df.fillna("").values:
                ws6.append(list(row))
            apply_table_style(ws6,start_row=1,start_col=1,max_row=ws6.max_row,max_col=ws6.shape[1] if hasattr(ws6,"shape") else len(raw_df.columns))

            output = BytesIO()
            wb.save(output)
            return output.getvalue()

        st.download_button(
            "📥 Download Premium MIS Excel",
            data=to_excel(final_df, df_raw),
            file_name="MCF_Premium_MIS.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error("❌ Error occurred")
        st.exception(e)
else:
    st.info("👆 Upload file to start")
